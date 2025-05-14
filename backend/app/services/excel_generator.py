from openpyxl import load_workbook

def fill_excel_generic(fields, template_path, output_path):
    wb = load_workbook(template_path)
    ws = wb.active

    for field in fields:
        name = field.get("name")
        value = field.get("value")

        if not name or value is None:
            continue

        # Try to find a cell with the same header and write the value in adjacent column
        for row in ws.iter_rows(min_row=1, max_row=100, max_col=10):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and name.lower() in cell.value.lower():
                    adjacent_cell = ws.cell(row=cell.row, column=cell.column + 1)
                    adjacent_cell.value = value
                    break  # Stop after finding first match

    wb.save(output_path)
