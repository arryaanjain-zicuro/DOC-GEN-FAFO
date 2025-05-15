# app/parsing/base/excel_utils.py
import openpyxl
from typing import Dict, List

def extract_excel_content(excel_file: str) -> Dict[str, List[List[str]]]:
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    content = {}
    for sheet in wb.sheetnames:
        rows = [
            [str(cell).strip() if cell is not None else "" for cell in row]
            for row in wb[sheet].iter_rows(values_only=True)
            if any(cell is not None for cell in row)
        ]
        content[sheet] = rows
    return content
