import os, re, json, time
import openpyxl
from dotenv import load_dotenv
from typing import Dict, Any, List
import google.generativeai as genai
from app.core.gemini_rate_limiter import GeminiRateLimiter

load_dotenv()

# Initialize Gemini API
genai.configure(api_key=os.getenv("GEMINI_API_KEY"))

model = genai.GenerativeModel("gemini-2.0-flash")

# Set rate-limiting parameters
CALLS = 20
RATE_LIMIT = 60  # seconds

# Custom rate limiter for Gemini
rate_limiter = GeminiRateLimiter(rate_limit_per_minute=CALLS)

def send_to_gemini(prompt: str, retries: int = 5) -> str:
    """Send request to Gemini with rate limiting."""
    for attempt in range(retries):
        try:
            rate_limiter.wait()  # Throttle before sending request
            response = model.generate_content(prompt)
            return response.text
        except Exception as e:
            print(f"[Attempt {attempt + 1}] Gemini error: {e}")
            if attempt == retries - 1:
                raise
            time.sleep(2)  # fixed delay between retries

def extract_excel_content(excel_file: str) -> Dict[str, List[List[str]]]:
    """Extract content from all sheets of an Excel file."""
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    content = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = []
        for row in ws.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                rows.append([str(cell).strip() if cell is not None else "" for cell in row])
        content[sheet] = rows
    return content

def build_comparison_prompt(alpha_analysis: Dict[str, Any], beta_excel_data: Dict[str, Any]) -> str:
    """Construct the comparison prompt for Gemini."""
    alpha_fields = json.dumps(alpha_analysis['fields'], indent=2)
    beta_raw = json.dumps(beta_excel_data, indent=2)
    return f"""
You are an AI system comparing ALPHA fields with a BETA Excel sheet.

Alpha (base template) fields:
{alpha_fields}

Beta (Excel) content:
{beta_raw}

Your tasks:
1. Identify which ALPHA fields are reused in BETA Excel.
2. Explain any transformations (math, formatting, references).
3. Detect new fields in BETA.
4. Output JSON:
    - "field_mappings": List of mappings {{"alpha_field": ..., "sheet": ..., "cell": ..., "value": ..., "transformation": ...}}
    - "unmatched_beta_cells": List of unlinked cells
    - "transformation_notes": Summary of any formulas or logic
"""

def analyze_excel_against_alpha(alpha_analysis: Dict[str, Any], beta_excel_data: Dict[str, Any]) -> Dict[str, Any]:
    """Send Gemini prompt with alpha fields and beta Excel data to analyze patterns."""
    prompt = build_comparison_prompt(alpha_analysis, beta_excel_data)
    gemini_response = send_to_gemini(prompt)
    cleaned_response = re.sub(r"^```json\s*|\s*```$", "", gemini_response.strip())

    # Now safely parse the JSON
    try:
        parsed_result = json.loads(cleaned_response)
    except json.JSONDecodeError as e:
        raise ValueError(f"Invalid JSON in Gemini response: {e}")
    
    return parsed_result

def parse_beta_excel(excel_path: str, alpha_analysis: Dict[str, Any]) -> Dict[str, Any]:
    """End-to-end parse of beta Excel document and comparison with alpha."""
    beta_data = extract_excel_content(excel_path)
    mapping_result = analyze_excel_against_alpha(alpha_analysis, beta_data)

    return {
        "beta_raw": beta_data,
        "mapping_result": mapping_result
    }

# Example usage
# if __name__ == "__main__":
#     import sys
#     from backend.app.parser.alpha_doc_parser import parse_alpha_document

#     if len(sys.argv) != 3:
#         print("Usage: python beta_excel_parser.py alpha.docx beta.xlsx")
#         sys.exit(1)

#     alpha_path = sys.argv[1]
#     beta_excel_path = sys.argv[2]

#     alpha = parse_alpha_document(alpha_path)["gemini_analysis"]
#     result = parse_beta_excel(beta_excel_path, alpha)

#     print("\n=== Excel Mappings ===")
#     print(json.dumps(result["mapping_result"], indent=2))
